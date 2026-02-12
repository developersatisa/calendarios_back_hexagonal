from typing import Optional, List, Dict, Any
from datetime import date, time, datetime
import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    Workbook = None

from app.infrastructure.db.database import SessionLocal
from app.infrastructure.db.models.cliente_proceso_hito_model import ClienteProcesoHitoModel
from app.infrastructure.db.models.cliente_proceso_model import ClienteProcesoModel
from app.infrastructure.db.models.cliente_model import ClienteModel
from app.infrastructure.db.models.proceso_model import ProcesoModel
from app.infrastructure.db.models.hito_model import HitoModel
from app.infrastructure.db.models.cliente_proceso_hito_cumplimiento_model import ClienteProcesoHitoCumplimientoModel
from app.infrastructure.db.models.documentos_cumplimiento_model import DocumentoCumplimientoModel

router = APIRouter(prefix="/status-todos-clientes", tags=["Status Todos los Clientes"])

# — Dependencias —

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# — Helpers de Consulta —

def _get_subquery_ultimo_cumplimiento(db: Session):
    """
    Subconsulta para obtener el ID del último cumplimiento por hito.
    """
    return (
        db.query(
            ClienteProcesoHitoCumplimientoModel.cliente_proceso_hito_id,
            func.max(ClienteProcesoHitoCumplimientoModel.id).label('ultimo_cumplimiento_id')
        )
        .group_by(ClienteProcesoHitoCumplimientoModel.cliente_proceso_hito_id)
        .subquery()
    )

def _build_base_query(db: Session, subquery_ultimo_cumplimiento):
    """
    Construye la consulta base con todos los JOINs necesarios.
    """
    return (
        db.query(
            # Campos de ClienteProcesoHito
            ClienteProcesoHitoModel.id,
            ClienteProcesoHitoModel.cliente_proceso_id,
            ClienteProcesoHitoModel.hito_id,
            ClienteProcesoHitoModel.estado,
            ClienteProcesoHitoModel.fecha_estado,
            ClienteProcesoHitoModel.fecha_limite,
            ClienteProcesoHitoModel.hora_limite,
            ClienteProcesoHitoModel.tipo,
            ClienteProcesoHitoModel.habilitado,

            # Información del cliente
            ClienteModel.idcliente.label('cliente_id'),
            ClienteModel.razsoc.label('cliente_nombre'),

            # Información del proceso
            ClienteProcesoModel.proceso_id,
            ProcesoModel.nombre.label('proceso_nombre'),

            # Información del hito maestro
            HitoModel.nombre.label('hito_nombre'),
            HitoModel.obligatorio.label('hito_obligatorio'),

            # Último cumplimiento (si existe)
            ClienteProcesoHitoCumplimientoModel.id.label('cumplimiento_id'),
            ClienteProcesoHitoCumplimientoModel.fecha.label('cumplimiento_fecha'),
            ClienteProcesoHitoCumplimientoModel.hora.label('cumplimiento_hora'),
            ClienteProcesoHitoCumplimientoModel.observacion.label('cumplimiento_observacion'),
            ClienteProcesoHitoCumplimientoModel.usuario.label('cumplimiento_usuario'),
            ClienteProcesoHitoCumplimientoModel.codSubDepar.label('cumplimiento_codSubDepar'),
            ClienteProcesoHitoCumplimientoModel.fecha_creacion.label('cumplimiento_fecha_creacion'),

            # Número de documentos del último cumplimiento
            func.count(DocumentoCumplimientoModel.id).label('num_documentos')
        )
        .join(ClienteProcesoModel, ClienteProcesoHitoModel.cliente_proceso_id == ClienteProcesoModel.id)
        .join(ClienteModel, ClienteProcesoModel.cliente_id == ClienteModel.idcliente)
        .join(ProcesoModel, ClienteProcesoModel.proceso_id == ProcesoModel.id)
        .join(HitoModel, ClienteProcesoHitoModel.hito_id == HitoModel.id)
        .outerjoin(
            subquery_ultimo_cumplimiento,
            ClienteProcesoHitoModel.id == subquery_ultimo_cumplimiento.c.cliente_proceso_hito_id
        )
        .outerjoin(
            ClienteProcesoHitoCumplimientoModel,
            ClienteProcesoHitoCumplimientoModel.id == subquery_ultimo_cumplimiento.c.ultimo_cumplimiento_id
        )
        .outerjoin(
            DocumentoCumplimientoModel,
            ClienteProcesoHitoCumplimientoModel.id == DocumentoCumplimientoModel.cumplimiento_id
        )
        .filter(ClienteProcesoHitoModel.habilitado == True)
        .group_by(
            ClienteProcesoHitoModel.id,
            ClienteProcesoHitoModel.cliente_proceso_id,
            ClienteProcesoHitoModel.hito_id,
            ClienteProcesoHitoModel.estado,
            ClienteProcesoHitoModel.fecha_estado,
            ClienteProcesoHitoModel.fecha_limite,
            ClienteProcesoHitoModel.hora_limite,
            ClienteProcesoHitoModel.tipo,
            ClienteProcesoHitoModel.habilitado,
            ClienteModel.idcliente,
            ClienteModel.razsoc,
            ClienteProcesoModel.proceso_id,
            ProcesoModel.nombre,
            HitoModel.nombre,
            HitoModel.obligatorio,
            ClienteProcesoHitoCumplimientoModel.id,
            ClienteProcesoHitoCumplimientoModel.fecha,
            ClienteProcesoHitoCumplimientoModel.hora,
            ClienteProcesoHitoCumplimientoModel.observacion,
            ClienteProcesoHitoCumplimientoModel.usuario,
            ClienteProcesoHitoCumplimientoModel.codSubDepar,
            ClienteProcesoHitoCumplimientoModel.fecha_creacion
        )
    )

def _apply_filters(
    query,
    fecha_limite_desde: Optional[str] = None,
    fecha_limite_hasta: Optional[str] = None,
    cliente_id: Optional[str] = None,
    proceso_id: Optional[int] = None,
    hito_id: Optional[int] = None
):
    """
    Aplica los filtros comunes a la consulta.
    """
    if fecha_limite_desde:
        try:
            fecha_desde = date.fromisoformat(fecha_limite_desde)
            query = query.filter(ClienteProcesoHitoModel.fecha_limite >= fecha_desde)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha_limite_desde inválido. Use YYYY-MM-DD")

    if fecha_limite_hasta:
        try:
            fecha_hasta = date.fromisoformat(fecha_limite_hasta)
            query = query.filter(ClienteProcesoHitoModel.fecha_limite <= fecha_hasta)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha_limite_hasta inválido. Use YYYY-MM-DD")

    if cliente_id:
        query = query.filter(ClienteModel.idcliente == cliente_id)

    if proceso_id:
        query = query.filter(ClienteProcesoModel.proceso_id == proceso_id)

    if hito_id:
        query = query.filter(ClienteProcesoHitoModel.hito_id == hito_id)

    return query

def _calculate_excel_status(estado_base, fecha_limite, hora_limite, fecha_cumplimiento):
    """
    Calcula el estado para el reporte Excel basado en reglas de negocio.
    """
    if estado_base == 'Finalizado':
        if not fecha_cumplimiento:
            return "Finalizado"

        # Construir deadline datetime
        if not fecha_limite:
            return "Finalizado"

        deadline = datetime.combine(fecha_limite, hora_limite) if hora_limite else datetime.combine(fecha_limite, time(23, 59, 59))

        # Normalizar fecha_cumplimiento a datetime
        if isinstance(fecha_cumplimiento, date) and not isinstance(fecha_cumplimiento, datetime):
            fulfillment_dt = datetime.combine(fecha_cumplimiento, time(0, 0, 0))
        else:
            fulfillment_dt = fecha_cumplimiento

        # Comparar
        if fulfillment_dt and fulfillment_dt > deadline:
            return "Cumplido fuera de plazo"
        else:
            return "Cumplido en plazo"

    else:
        # Estados pendientes
        if not fecha_limite:
            return estado_base

        today = date.today()

        if fecha_limite == today:
            return "Vence hoy"
        elif fecha_limite < today:
            return "Pendiente fuera de plazo"
        else:
            return "Pendiente en plazo"

# — Endpoints —

@router.get("/hitos", summary="Obtener todos los hitos habilitados de todos los clientes",
            description="Devuelve todos los hitos habilitados de todos los clientes con información relacionada (cliente, proceso, hito maestro) y el último cumplimiento si existe.")
def get_status_todos_clientes(
    fecha_limite_desde: Optional[str] = Query(None, description="Filtrar por fecha límite desde (YYYY-MM-DD)"),
    fecha_limite_hasta: Optional[str] = Query(None, description="Filtrar por fecha límite hasta (YYYY-MM-DD)"),
    cliente_id: Optional[str] = Query(None, description="Filtrar por ID de cliente"),
    proceso_id: Optional[int] = Query(None, description="Filtrar por ID de proceso"),
    hito_id: Optional[int] = Query(None, description="Filtrar por ID de hito"),
    ordenar_por: Optional[str] = Query("fecha_limite", description="Campo para ordenar (fecha_limite, cliente_nombre, proceso_nombre)"),
    orden: Optional[str] = Query("asc", description="Orden (asc o desc)"),
    limit: Optional[int] = Query(None, ge=1, le=10000, description="Límite de resultados"),
    offset: Optional[int] = Query(None, ge=0, description="Offset para paginación"),
    db: Session = Depends(get_db)
):
    try:
        # 1. Preparar Query
        subquery = _get_subquery_ultimo_cumplimiento(db)
        query = _build_base_query(db, subquery)

        # 2. Aplicar Filtros
        query = _apply_filters(query, fecha_limite_desde, fecha_limite_hasta, cliente_id, proceso_id, hito_id)

        # 3. Aplicar Ordenamiento
        if ordenar_por == "fecha_limite":
            order_field = ClienteProcesoHitoModel.fecha_limite
        elif ordenar_por == "cliente_nombre":
            order_field = ClienteModel.razsoc
        elif ordenar_por == "proceso_nombre":
            order_field = ProcesoModel.nombre
        else:
            order_field = ClienteProcesoHitoModel.fecha_limite

        if orden.lower() == "desc":
            query = query.order_by(order_field.desc())
        else:
            query = query.order_by(order_field.asc())

        # 4. Obtener Total (antes de paginar)
        total = query.count()

        # 5. Aplicar Paginación
        if offset is not None:
            query = query.offset(offset)
        if limit is not None:
            query = query.limit(limit)

        # 6. Ejecutar
        resultados = query.all()

        # 7. Mapear Respuesta
        hitos_response = []
        for row in resultados:
            ultimo_cumplimiento = None
            if row.cumplimiento_id:
                ultimo_cumplimiento = {
                    "id": row.cumplimiento_id,
                    "fecha": row.cumplimiento_fecha.isoformat() if row.cumplimiento_fecha else None,
                    "hora": str(row.cumplimiento_hora) if row.cumplimiento_hora else None,
                    "observacion": row.cumplimiento_observacion,
                    "usuario": row.cumplimiento_usuario,
                    "fecha_creacion": row.cumplimiento_fecha_creacion.isoformat() if row.cumplimiento_fecha_creacion else None,
                    "num_documentos": int(row.num_documentos or 0)
                }

            hito_dict = {
                "id": row.id,
                "cliente_proceso_id": row.cliente_proceso_id,
                "hito_id": row.hito_id,
                "estado": row.estado,
                "fecha_estado": row.fecha_estado.isoformat() if row.fecha_estado else None,
                "fecha_limite": row.fecha_limite.isoformat() if row.fecha_limite else None,
                "hora_limite": str(row.hora_limite) if row.hora_limite else None,
                "tipo": row.tipo,
                "habilitado": bool(row.habilitado),
                "cliente_id": str(row.cliente_id or ""),
                "cliente_nombre": str(row.cliente_nombre or "").strip(),
                "proceso_id": row.proceso_id,
                "proceso_nombre": str(row.proceso_nombre or "").strip(),
                "hito_nombre": str(row.hito_nombre or "").strip(),
                "ultimo_cumplimiento": ultimo_cumplimiento
            }
            hitos_response.append(hito_dict)

        return {
            "hitos": hitos_response,
            "total": total
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener hitos: {str(e)}")


@router.get("/exportar-excel", summary="Exportar status de todos los clientes a Excel",
            description="Genera y descarga un archivo Excel con el estado de los hitos filtrados, utilizando colores para indicar el estado de cumplimiento.")
def exportar_status_todos_excel(
    fecha_limite_desde: Optional[str] = Query(None, alias="fecha_desde", description="Filtrar por fecha límite desde (YYYY-MM-DD)"),
    fecha_limite_hasta: Optional[str] = Query(None, alias="fecha_hasta", description="Filtrar por fecha límite hasta (YYYY-MM-DD)"),
    cliente_id: Optional[str] = Query(None, description="Filtrar por ID de cliente"),
    proceso_nombre: Optional[str] = Query(None, description="Filtrar por nombre del proceso"),
    hito_id: Optional[int] = Query(None, description="Filtrar por ID de hito"),
    estados: Optional[str] = Query(None, description="Filtrar por estados (separados por coma): cumplido_en_plazo,cumplido_fuera_plazo,vence_hoy,pendiente_fuera_plazo,pendiente_en_plazo"),
    tipos: Optional[str] = Query(None, description="Filtrar por tipos (separados por coma): Atisa,Cliente,Terceros"),
    search_term: Optional[str] = Query(None, description="Búsqueda de texto libre en proceso_nombre y hito_nombre"),
    db: Session = Depends(get_db)
):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="La librería 'openpyxl' no está instalada.")

    try:
        # 1. Preparar Query
        subquery = _get_subquery_ultimo_cumplimiento(db)
        query = _build_base_query(db, subquery)

        # 2. Aplicar filtros básicos
        query = _apply_filters(query, fecha_limite_desde, fecha_limite_hasta, cliente_id, None, hito_id)

        # 3. Filtrar por nombre de proceso (búsqueda parcial)
        if proceso_nombre:
            query = query.filter(ProcesoModel.nombre.ilike(f"%{proceso_nombre}%"))

        # 4. Filtrar por tipos
        if tipos:
            tipos_list = [t.strip() for t in tipos.split(",")]
            query = query.filter(ClienteProcesoHitoModel.tipo.in_(tipos_list))

        # 5. Búsqueda de texto libre
        if search_term:
            search_pattern = f"%{search_term}%"
            query = query.filter(
                (ProcesoModel.nombre.ilike(search_pattern)) |
                (HitoModel.nombre.ilike(search_pattern))
            )

        # 6. Ordenar
        query = query.order_by(ClienteProcesoHitoModel.fecha_limite.asc())

        # 7. Ejecutar
        resultados = query.all()

        # 8. Filtrar por estados calculados (post-query)
        if estados:
            estados_list = [e.strip() for e in estados.split(",")]
            filtered_results = []
            for r in resultados:
                estado_calc = _calculate_excel_status(r.estado, r.fecha_limite, r.hora_limite, r.cumplimiento_fecha)
                # Normalizar a snake_case
                estado_key = estado_calc.lower().replace(" ", "_")
                if estado_key in estados_list:
                    filtered_results.append(r)
            resultados = filtered_results

        # 9. Generar Excel con dos hojas
        wb = Workbook()

        # PRIMERA HOJA: Filtros Aplicados
        ws_filtros = wb.active
        ws_filtros.title = "Filtros Aplicados"

        # Título
        ws_filtros.append(["FILTROS APLICADOS"])
        ws_filtros["A1"].font = Font(size=14, bold=True)
        ws_filtros.append(["-" * 50])
        ws_filtros.append([])

        # Detalles de filtros
        ws_filtros.append(["Cliente:", cliente_id or "Todos"])
        ws_filtros.append(["Proceso:", proceso_nombre or "Todos"])
        ws_filtros.append(["Hito ID:", str(hito_id) if hito_id else "Todos"])
        ws_filtros.append(["Fecha Desde:", fecha_limite_desde or "Sin filtro"])
        ws_filtros.append(["Fecha Hasta:", fecha_limite_hasta or "Sin filtro"])
        ws_filtros.append(["Estados:", estados.replace(",", ", ") if estados else "Todos"])
        ws_filtros.append(["Tipos:", tipos.replace(",", ", ") if tipos else "Todos"])
        ws_filtros.append(["Búsqueda:", search_term or "Sin búsqueda"])
        ws_filtros.append([])
        ws_filtros.append(["Fecha de Generación:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
        ws_filtros.append(["Total de Registros:", len(resultados)])

        # Ajustar ancho de columnas
        ws_filtros.column_dimensions["A"].width = 25
        ws_filtros.column_dimensions["B"].width = 50

        # SEGUNDA HOJA: Datos
        ws_datos = wb.create_sheet("Datos")

        headers = ["Cliente", "Proceso", "Hito", "Fecha Límite", "Hora Límite", "Estado",
                   "Fecha Estado", "Tipo", "Obligatorio", "Observaciones"]
        ws_datos.append(headers)

        # Estilo headers
        for cell in ws_datos[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Colores por estado
        colores_estado = {
            "Cumplido en plazo": PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid"),
            "Cumplido fuera de plazo": PatternFill(start_color="b45309", end_color="b45309", fill_type="solid"),
            "Vence hoy": PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid"),
            "Pendiente fuera de plazo": PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid"),
            "Pendiente en plazo": PatternFill(start_color="00a1de", end_color="00a1de", fill_type="solid"),
        }
        font_blanco = Font(color="FFFFFF", bold=False)

        # Llenar filas
        for r in resultados:
            estado_calculado = _calculate_excel_status(r.estado, r.fecha_limite, r.hora_limite, r.cumplimiento_fecha)

            ws_datos.append([
                str(r.cliente_nombre or "").strip(),
                str(r.proceso_nombre or "").strip(),
                str(r.hito_nombre or "").strip(),
                r.fecha_limite.strftime("%d/%m/%Y") if r.fecha_limite else "",
                r.hora_limite.strftime("%H:%M") if r.hora_limite else "",
                estado_calculado,
                r.fecha_estado.strftime("%d/%m/%Y %H:%M") if r.fecha_estado else "",
                r.tipo,
                "Sí" if getattr(r, 'hito_obligatorio', 0) == 1 else "No",
                str(r.cumplimiento_observacion or "").strip() if r.cumplimiento_id else ""
            ])

            # Aplicar color
            fila_numero = ws_datos.max_row
            fill_color = colores_estado.get(estado_calculado)
            if fill_color:
                for cell in ws_datos[fila_numero]:
                    cell.fill = fill_color
                    cell.font = font_blanco

        # Auto-ajustar columnas
        for col in ws_datos.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value)
                    if len(val) > max_length:
                        max_length = len(val)
                except:
                    pass
            ws_datos.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"status_todos_clientes_{date.today().strftime('%Y-%m-%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename={filename}',
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar Excel: {str(e)}")
