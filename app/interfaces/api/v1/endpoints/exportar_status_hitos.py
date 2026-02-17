# app/interfaces/api/v1/endpoints/exportar_status_hitos.py

from fastapi import APIRouter, Depends, HTTPException, Query, Path
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date, time
from typing import Optional, List
import io

from app.infrastructure.db.database import SessionLocal
from app.infrastructure.db.repositories.cliente_proceso_repository_sql import ClienteProcesoRepositorySQL
from app.infrastructure.db.repositories.cliente_proceso_hito_repository_sql import ClienteProcesoHitoRepositorySQL
from app.infrastructure.db.repositories.cliente_proceso_hito_cumplimiento_repository_sql import ClienteProcesoHitoCumplimientoRepositorySQL
from app.infrastructure.db.models.cliente_proceso_hito_model import ClienteProcesoHitoModel
from app.infrastructure.db.models.cliente_proceso_model import ClienteProcesoModel
from app.infrastructure.db.models.proceso_model import ProcesoModel
from app.infrastructure.db.models.hito_model import HitoModel
from app.infrastructure.db.models.cliente_proceso_hito_cumplimiento_model import ClienteProcesoHitoCumplimientoModel
from app.infrastructure.db.models.proceso_hito_maestro_model import ProcesoHitoMaestroModel
from sqlalchemy import case
from datetime import datetime, time as dt_time

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise ImportError("openpyxl no está instalado. Instálalo con: pip install openpyxl")

router = APIRouter(prefix="/status-cliente", tags=["Exportar Status Hitos"])

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_repo_cp(db: Session = Depends(get_db)):
    return ClienteProcesoRepositorySQL(db)

def get_repo_cph(db: Session = Depends(get_db)):
    return ClienteProcesoHitoRepositorySQL(db)

def get_repo_cumplimiento(db: Session = Depends(get_db)):
    return ClienteProcesoHitoCumplimientoRepositorySQL(db)

def calcular_estado_hito(
    estado: str,
    fecha_limite: Optional[date],
    hora_limite: Optional[time],
    cumplimientos: List,
    fecha_actual: date = None
) -> str:
    """
    Calcula el estado del hito según la lógica unificada:
    1. Si estado = 'Finalizado':
       - Compara fecha/hora ultimo cumplimiento con fecha/hora limite.
    2. Si no es 'Finalizado':
       - Compara fecha limite con hoy.
    """

    # 1. Obtener fecha/hora del último cumplimiento si existe
    fecha_cumplimiento = None
    hora_cumplimiento = None

    if cumplimientos:
        # Busca el cumplimiento más reciente
        ultimo = max(cumplimientos, key=lambda c: (c.fecha, c.hora or dt_time.min))
        fecha_cumplimiento = ultimo.fecha
        hora_cumplimiento = ultimo.hora

    if estado == 'Finalizado':
        if not fecha_cumplimiento:
            # Si está finalizado pero no tiene cumplimiento registrado (caso borde),
            # devolvemos "Finalizado" o asumimos plazo.
            # Según lógica anterior: "Finalizado" es neutro si falta info,
            # pero el frontend suele poner "Cumplido en plazo" por defecto si no falla.
            # Usaremos "Finalizado" para ser consistentes con status_todos_clientes
            return "Finalizado"

        if not fecha_limite:
            return "Finalizado"

        # Construir datetimes para comparación precisa
        # Si no hay hora límite, asumimos final del día
        deadline = datetime.combine(fecha_limite, hora_limite) if hora_limite else datetime.combine(fecha_limite, dt_time(23, 59, 59))

        # Si no hay hora cumplimiento, asumimos inicio del día? O lo que venga.
        # status_todos_clientes usaba time(0,0,0) si fecha_cumplimiento era solo date (aunque allí venía de DB join).
        # Aquí viene de objeto ORM.
        fulfillment_dt = datetime.combine(fecha_cumplimiento, hora_cumplimiento if hora_cumplimiento else dt_time(0, 0, 0))

        if fulfillment_dt > deadline:
            return "Cumplido fuera de plazo"
        else:
            return "Cumplido en plazo"

    else:
        # Estado Pendiente, Nuevo, En Progreso, etc.
        if not fecha_limite:
            return estado

        today = date.today()

        if fecha_limite == today:
            return "Vence hoy"
        elif fecha_limite < today:
            return "Pendiente fuera de plazo"
        else:
            return "Pendiente en plazo"

def formatear_fecha(fecha: Optional[date]) -> str:
    """Formatea una fecha a DD/MM/YYYY"""
    if fecha is None:
        return ""
    return fecha.strftime("%d/%m/%Y")

def formatear_hora(hora: Optional[time]) -> str:
    """Formatea una hora a HH:MM"""
    if hora is None:
        return ""
    return hora.strftime("%H:%M")

@router.get("/{cliente_id}/exportar-excel")
def exportar_status_hitos_excel(
    cliente_id: str = Path(..., description="ID del cliente"),
    hito_id: Optional[int] = Query(None, description="Filtrar por ID de hito específico"),
    proceso_nombre: Optional[str] = Query(None, description="Filtrar por nombre de proceso"),
    fecha_desde: Optional[str] = Query(None, description="Fecha límite desde (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha límite hasta (YYYY-MM-DD)"),
    estados: Optional[str] = Query(None, description="Estados separados por comas"),
    tipos: Optional[str] = Query(None, description="Tipos de hito separados por comas"),
    search_term: Optional[str] = Query(None, description="Búsqueda por texto"),
    db: Session = Depends(get_db)
):
    """
    Exporta a Excel los hitos de un cliente con los filtros aplicados.

    IMPORTANTE: Este endpoint exporta TODOS los hitos que cumplan con los filtros,
    sin límites de paginación. Todos los resultados se incluyen en el archivo Excel.
    """
    try:
        # Parsear fechas
        fecha_desde_parsed = None
        fecha_hasta_parsed = None
        if fecha_desde:
            fecha_desde_parsed = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
        if fecha_hasta:
            fecha_hasta_parsed = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()

        # Parsear estados y tipos
        # Mapeo de estados con guiones bajos a estados con espacios (como se calculan)
        estado_mapping = {
            "cumplido_en_plazo": "Cumplido en plazo",
            "cumplido_fuera_plazo": "Cumplido fuera de plazo",
            "vence_hoy": "Vence hoy",
            "pendiente_fuera_plazo": "Pendiente fuera de plazo",
            "pendiente_en_plazo": "Pendiente en plazo"
        }

        estados_raw = [e.strip() for e in estados.split(",")] if estados else []
        # Convertir estados con guiones bajos a estados con espacios
        estados_list = [estado_mapping.get(estado.lower(), estado) for estado in estados_raw]
        tipos_list = [t.strip() for t in tipos.split(",")] if tipos else []

        # Obtener procesos habilitados del cliente
        repo_cp = ClienteProcesoRepositorySQL(db)
        procesos_cliente = repo_cp.listar_habilitados_por_cliente(cliente_id)

        if not procesos_cliente:
            raise HTTPException(status_code=404, detail=f"No se encontraron procesos habilitados para el cliente {cliente_id}")

        # Obtener todos los hitos habilitados de los procesos del cliente usando una query optimizada

        # Construir lista de IDs de procesos del cliente
        proceso_ids = [cp.id for cp in procesos_cliente]
        if not proceso_ids:
            raise HTTPException(status_code=404, detail=f"No se encontraron procesos habilitados para el cliente {cliente_id}")

        # Query optimizada para obtener todos los hitos con sus relaciones
        query = db.query(
            ClienteProcesoHitoModel,
            ProcesoModel,
            HitoModel
        ).join(
            ClienteProcesoModel, ClienteProcesoHitoModel.cliente_proceso_id == ClienteProcesoModel.id
        ).join(
            ProcesoModel, ClienteProcesoModel.proceso_id == ProcesoModel.id
        ).join(
            ProcesoHitoMaestroModel, ClienteProcesoHitoModel.hito_id == ProcesoHitoMaestroModel.hito_id
        ).join(
            HitoModel, ProcesoHitoMaestroModel.hito_id == HitoModel.id
        ).filter(
            ClienteProcesoHitoModel.cliente_proceso_id.in_(proceso_ids),
            ClienteProcesoHitoModel.habilitado == True
        )

        # Aplicar filtros adicionales
        if hito_id:
            query = query.filter(ClienteProcesoHitoModel.hito_id == hito_id)

        if proceso_nombre:
            query = query.filter(ProcesoModel.nombre.ilike(f"%{proceso_nombre}%"))

        if fecha_desde_parsed:
            query = query.filter(ClienteProcesoHitoModel.fecha_limite >= fecha_desde_parsed)

        if fecha_hasta_parsed:
            query = query.filter(ClienteProcesoHitoModel.fecha_limite <= fecha_hasta_parsed)

        if tipos_list:
            query = query.filter(ClienteProcesoHitoModel.tipo.in_(tipos_list))

        # Ordenar por fecha límite y hora límite ascendente
        # SQL Server no soporta NULLS LAST, usamos CASE para poner NULLs al final
        query = query.order_by(
            case(
                (ClienteProcesoHitoModel.fecha_limite.is_(None), 1),
                else_=0
            ),
            ClienteProcesoHitoModel.fecha_limite.asc(),
            case(
                (ClienteProcesoHitoModel.hora_limite.is_(None), 1),
                else_=0
            ),
            ClienteProcesoHitoModel.hora_limite.asc()
        )

        # Obtener TODOS los resultados sin límite de paginación
        # Usar yield_per para procesar en lotes si hay muchos registros, pero obtener todos
        resultados = query.all()

        # Obtener cumplimientos para todos los hitos de una vez
        hito_ids = [r[0].id for r in resultados]
        cumplimientos_dict = {}
        if hito_ids:
            cumplimientos = db.query(ClienteProcesoHitoCumplimientoModel).filter(
                ClienteProcesoHitoCumplimientoModel.cliente_proceso_hito_id.in_(hito_ids)
            ).all()

            for cumplimiento in cumplimientos:
                if cumplimiento.cliente_proceso_hito_id not in cumplimientos_dict:
                    cumplimientos_dict[cumplimiento.cliente_proceso_hito_id] = []
                cumplimientos_dict[cumplimiento.cliente_proceso_hito_id].append(cumplimiento)

        # Procesar resultados y aplicar filtros de estado y búsqueda
        hitos_data = []

        for hito_model, proceso_model, hito_maestro_model in resultados:
            # Obtener cumplimientos del hito
            cumplimientos = cumplimientos_dict.get(hito_model.id, [])

            # Calcular estado
            estado_calculado = calcular_estado_hito(
                hito_model.estado,
                hito_model.fecha_limite,
                hito_model.hora_limite,
                cumplimientos
            )

            # Filtrar por estados calculados
            if estados_list and estado_calculado not in estados_list:
                continue

            # Filtrar por search_term
            if search_term:
                search_lower = search_term.lower()
                if not (
                    search_lower in proceso_model.nombre.lower() or
                    search_lower in hito_maestro_model.nombre.lower() or
                    search_lower in estado_calculado.lower() or
                    search_lower in hito_model.tipo.lower()
                ):
                    continue

            # Agregar a la lista
            hitos_data.append({
                'proceso_nombre': proceso_model.nombre,
                'hito_nombre': hito_maestro_model.nombre,
                'estado_calculado': estado_calculado,
                'fecha_limite': hito_model.fecha_limite,
                'hora_limite': hito_model.hora_limite,
                'fecha_estado': hito_model.fecha_estado.date() if hito_model.fecha_estado else None,
                'tipo': hito_model.tipo
            })

        if not hitos_data:
            raise HTTPException(status_code=404, detail="No se encontraron hitos que cumplan con los filtros especificados")

        # Crear el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Status de Hitos"

        # Encabezados
        headers = ["Proceso", "Hito", "Estado", "Fecha Límite", "Hora Límite", "Fecha Estado", "Tipo"]
        ws.append(headers)

        # Formatear encabezados en negrita
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Definir colores de fondo por estado (según los colores del frontend)
        colores_estado = {
            "Cumplido en plazo": PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid"),  # Verde
            "Cumplido fuera de plazo": PatternFill(start_color="b45309", end_color="b45309", fill_type="solid"),  # Naranja
            "Vence hoy": PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid"),  # Rojo
            "Pendiente fuera de plazo": PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid"),  # Rojo claro
            "Pendiente en plazo": PatternFill(start_color="00a1de", end_color="00a1de", fill_type="solid"),  # Azul Atisa
        }

        # Color de texto blanco para todos los estados
        font_blanco = Font(color="FFFFFF", bold=False)

        # Agregar datos con estilos
        for hito_data in hitos_data:
            estado = hito_data['estado_calculado']
            fila_numero = ws.max_row + 1

            # Agregar la fila
            ws.append([
                hito_data['proceso_nombre'],
                hito_data['hito_nombre'],
                estado,
                formatear_fecha(hito_data['fecha_limite']),
                formatear_hora(hito_data['hora_limite']),
                formatear_fecha(hito_data['fecha_estado']),
                hito_data['tipo']
            ])

            # Aplicar color de fondo y texto blanco a toda la fila
            fill_color = colores_estado.get(estado, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))
            for cell in ws[fila_numero]:
                cell.fill = fill_color
                cell.font = font_blanco
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generar archivo en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Nombre del archivo
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        filename = f"status_hitos_cliente_{cliente_id}_{fecha_actual}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename={filename}',
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Error en el formato de fecha: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar el archivo Excel: {str(e)}")
