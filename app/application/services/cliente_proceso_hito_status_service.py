from typing import Optional, List, Dict, Any
from datetime import date, time, datetime
from calendar import monthrange
import io
from fastapi import HTTPException

from app.domain.repositories.cliente_proceso_hito_repository import ClienteProcesoHitoRepository

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    Workbook = None

class ClienteProcesoHitoStatusService:
    def __init__(self, repository: ClienteProcesoHitoRepository):
        self.repository = repository

    def _calculate_excel_status(self, estado_base, fecha_limite, hora_limite, fecha_cumplimiento):
        """
        Calcula el estado para el reporte Excel basado en reglas de negocio.
        """
        if estado_base == 'Finalizado':
            if not fecha_cumplimiento:
                return "Finalizado"

            # Construir deadline datetime
            if not fecha_limite:
                return "Finalizado"

            deadline = datetime.combine(fecha_limite, hora_limite) if hora_limite else datetime.combine(fecha_limite, time(23, 59, 59))

            # Normalizar fecha_cumplimiento a datetime
            if isinstance(fecha_cumplimiento, date) and not isinstance(fecha_cumplimiento, datetime):
                fulfillment_dt = datetime.combine(fecha_cumplimiento, time(0, 0, 0))
            else:
                fulfillment_dt = fecha_cumplimiento

            # Comparar
            if fulfillment_dt and fulfillment_dt > deadline:
                return "Cumplido fuera de plazo"
            else:
                return "Cumplido en plazo"

        else:
            # Estados pendientes
            if not fecha_limite:
                return estado_base

            today = date.today()

            if fecha_limite == today:
                return "Vence hoy"
            elif fecha_limite < today:
                return "Pendiente fuera de plazo"
            else:
                return "Pendiente en plazo"

    def obtener_reporte_status(self, filtros: dict, paginacion: dict):
        resultados, total = self.repository.ejecutar_reporte_status_todos_clientes(filtros, paginacion)

        hitos_response = []
        for row in resultados:
            ultimo_cumplimiento = None
            if row.cumplimiento_id:
                ultimo_cumplimiento = {
                    "id": row.cumplimiento_id,
                    "fecha": row.cumplimiento_fecha.isoformat() if row.cumplimiento_fecha else None,
                    "hora": str(row.cumplimiento_hora) if row.cumplimiento_hora else None,
                    "observacion": row.cumplimiento_observacion,
                    "usuario": row.cumplimiento_usuario,
                    "departamento": str(row.cumplimiento_departamento or "").strip(),
                    "codSubDepar": row.cumplimiento_codSubDepar,
                    "fecha_creacion": row.cumplimiento_fecha_creacion.isoformat() if row.cumplimiento_fecha_creacion else None,
                    "num_documentos": int(row.num_documentos or 0)
                }

            estado_calc = self._calculate_excel_status(row.estado, row.fecha_limite, row.hora_limite, row.cumplimiento_fecha)

            hito_dict = {
                "id": row.id,
                "cliente_proceso_id": row.cliente_proceso_id,
                "hito_id": row.hito_id,
                "estado": row.estado,
                "estado_calculado": estado_calc,
                "estado_proceso": getattr(row, 'proceso_estado', 'En proceso'),
                "fecha_estado": row.fecha_estado.isoformat() if row.fecha_estado else None,
                "fecha_limite": row.fecha_limite.isoformat() if row.fecha_limite else None,
                "hora_limite": str(row.hora_limite) if row.hora_limite else None,
                "tipo": row.tipo,
                "habilitado": bool(row.habilitado),
                "cliente_id": str(row.cliente_id or ""),
                "cliente_nombre": str(row.cliente_nombre or "").strip(),
                "codSubDepar": row.cliente_departamento_codigo,
                "departamento_cliente": str(getattr(row, 'cliente_departamento_nombre', '') or "").strip(),
                "proceso_id": row.proceso_id,
                "proceso_nombre": str(row.proceso_nombre or "").strip(),
                "hito_nombre": str(row.hito_nombre or "").strip(),
                "obligatorio": bool(getattr(row, 'hito_obligatorio', 0) == 1),
                "critico": bool(getattr(row, 'hito_critico', False)),
                "ultimo_cumplimiento": ultimo_cumplimiento
            }
            hitos_response.append(hito_dict)

        return {
            "hitos": hitos_response,
            "total": total
        }

    def exportar_reporte_excel(self, filtros: dict):
        if Workbook is None:
            raise HTTPException(status_code=500, detail="La librería 'openpyxl' no está instalada.")

        # 1. Obtener datos (sin paginación)
        resultados, _ = self.repository.ejecutar_reporte_status_todos_clientes(filtros, {})

        return self._generar_excel(resultados, filtros)

    def exportar_reporte_excel_por_usuario(self, filtros: dict, email: str):
        if Workbook is None:
            raise HTTPException(status_code=500, detail="La librería 'openpyxl' no está instalada.")

        # 1. Obtener datos filtrados por usuario (sin paginación)
        resultados, _ = self.repository.ejecutar_reporte_status_todos_clientes_por_usuario(filtros, {}, email)

        return self._generar_excel(resultados, filtros)

    def _generar_excel(self, resultados, filtros):
        # 2. Filtrar por estados calculados (post-query)
        estados = filtros.get('estados')
        if estados:
            estados_list = [e.strip() for e in estados.split(",")]
            filtered_results = []
            for r in resultados:
                estado_calc = self._calculate_excel_status(r.estado, r.fecha_limite, r.hora_limite, r.cumplimiento_fecha)
                # Normalizar a snake_case
                estado_key = estado_calc.lower().replace(" ", "_")
                if estado_key in estados_list:
                    filtered_results.append(r)
            resultados = filtered_results

        # 3. Generar Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Status"

        # --- SECCIÓN FILTROS ---
        ws.append(["FILTROS APLICADOS"])
        ws["A1"].font = Font(size=14, bold=True)
        ws.append(["-" * 50])
        ws.append([])

        # Detalles de filtros
        ws.append(["Cliente:", filtros.get('cliente_id') or "Todos"])
        ws.append(["Proceso:", filtros.get('proceso_nombre') or "Todos"])
        ws.append(["Hito ID:", str(filtros.get('hito_id')) if filtros.get('hito_id') else "Todos"])
        ws.append(["Fecha Desde:", filtros.get('fecha_limite_desde') or "Sin filtro"])
        ws.append(["Fecha Hasta:", filtros.get('fecha_limite_hasta') or "Sin filtro"])
        ws.append(["Estados:", estados.replace(",", ", ") if estados else "Todos"])
        ws.append(["Tipos:", filtros.get('tipos').replace(",", ", ") if filtros.get('tipos') else "Todos"])
        ws.append(["Búsqueda:", filtros.get('search_term') or "Sin búsqueda"])
        ws.append([])
        ws.append(["Fecha de Generación:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
        ws.append(["Total de Registros:", len(resultados)])
        ws.append([])
        ws.append([])

        # --- SECCIÓN DATOS ---
        headers = ["Cliente", "Cubo", "Proceso", "Periodo", "Estado Proceso", "Hito", "Responsable", "Clave",
                   "Estado", "Fecha Límite", "Hora Límite", "Fecha y Hora Actualización", "Gestor", "Observaciones",
                   "Fecha Cumplimiento", "Hora Cumplimiento",
                   "Fecha Creación Cumplimiento", "Cubo Cumplimiento"]
        ws.append(headers)

        # Estilo headers (fila actual es max_row)
        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Colores por estado
        colores_estado = {
            "Cumplido en plazo": PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid"),
            "Cumplido fuera de plazo": PatternFill(start_color="b45309", end_color="b45309", fill_type="solid"),
            "Vence hoy": PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid"),
            "Pendiente fuera de plazo": PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid"),
            "Pendiente en plazo": PatternFill(start_color="00a1de", end_color="00a1de", fill_type="solid"),
        }
        font_blanco = Font(color="FFFFFF", bold=False)

        # Agrupar resultados por cliente_proceso_id para calcular estado del proceso
        procesos_estado = {}
        for r in resultados:
            cp_id = r.cliente_proceso_id
            if cp_id not in procesos_estado:
                procesos_estado[cp_id] = {
                    'total_hitos': 0,
                    'hitos_finalizados': 0,
                    'fecha_inicio': r.proceso_fecha_inicio,
                    'fecha_fin': r.proceso_fecha_fin
                }
            procesos_estado[cp_id]['total_hitos'] += 1
            if r.estado == 'Finalizado':
                procesos_estado[cp_id]['hitos_finalizados'] += 1

        # Calcular estado de cada proceso
        for cp_id in procesos_estado:
            info = procesos_estado[cp_id]
            if info['hitos_finalizados'] == info['total_hitos']:
                info['estado'] = 'Finalizado'
            else:
                info['estado'] = 'En proceso'

        # Llenar filas
        for r in resultados:
            estado_calculado = self._calculate_excel_status(r.estado, r.fecha_limite, r.hora_limite, r.cumplimiento_fecha)

            # Obtener info del proceso
            cp_info = procesos_estado.get(r.cliente_proceso_id, {})
            periodo = ""

            # Priorizar mes y año para calcular periodo exacto
            if r.proceso_mes and r.proceso_anio:
                try:
                    _, last_day = monthrange(r.proceso_anio, r.proceso_mes)
                    inicio = date(r.proceso_anio, r.proceso_mes, 1)
                    fin = date(r.proceso_anio, r.proceso_mes, last_day)
                    periodo = f"{inicio.strftime('%d/%m/%Y')} - {fin.strftime('%d/%m/%Y')}"
                except ValueError:
                    # Fallback a fecha_inicio y fecha_fin
                    if r.proceso_fecha_inicio:
                        periodo = r.proceso_fecha_inicio.strftime("%d/%m/%Y")
                        if r.proceso_fecha_fin:
                            periodo += f" - {r.proceso_fecha_fin.strftime('%d/%m/%Y')}"
            # Fallback normal
            elif r.proceso_fecha_inicio:
                periodo = r.proceso_fecha_inicio.strftime("%d/%m/%Y")
                if r.proceso_fecha_fin:
                    periodo += f" - {r.proceso_fecha_fin.strftime('%d/%m/%Y')}"

            estado_proceso = cp_info.get('estado', 'En proceso')

            # Definir valores de cumplimiento si existen
            gestor = ""
            observaciones = ""
            fecha_creacion = ""
            fecha_cumplimiento = ""
            hora_cumplimiento = ""
            dept_cumplimiento = ""

            if getattr(r, 'cumplimiento_id', None):
                gestor = str(r.cumplimiento_usuario or "").strip()
                observaciones = str(r.cumplimiento_observacion or "").strip()
                fecha_creacion = r.cumplimiento_fecha_creacion.strftime("%d/%m/%Y %H:%M") if r.cumplimiento_fecha_creacion else ""
                fecha_cumplimiento = r.cumplimiento_fecha.strftime("%d/%m/%Y") if r.cumplimiento_fecha else ""
                hora_cumplimiento = r.cumplimiento_hora.strftime("%H:%M") if r.cumplimiento_hora else ""

                dept_cump_name = str(r.cumplimiento_departamento or "").strip()
                dept_cump_code = str(r.cumplimiento_codSubDepar or "").strip()
                if dept_cump_code:
                     suffix_cump = dept_cump_code[-2:] if len(dept_cump_code) >= 2 else dept_cump_code
                     dept_cumplimiento = f"{suffix_cump} - {dept_cump_name}"
                else:
                     dept_cumplimiento = dept_cump_name

            # Formatear Departamento Combinado
            dept_nombre = str(getattr(r, 'cliente_departamento_nombre', '') or "").strip()
            dept_codigo = str(getattr(r, 'cliente_departamento_codigo', '') or "").strip()
            dept_combined = ""
            if dept_codigo:
                # Tomar ultimos 2 digitos del codigo y concatenar con nombre
                suffix = dept_codigo[-2:] if len(dept_codigo) >= 2 else dept_codigo
                dept_combined = f"{suffix} - {dept_nombre}"
            else:
                dept_combined = dept_nombre

            ws.append([
                str(r.cliente_nombre or "").strip(),
                dept_combined, # Departamento (Cubo - Linea)
                str(r.proceso_nombre or "").strip(),
                periodo,
                estado_proceso,
                str(r.hito_nombre or "").strip(),
                str(r.tipo or ""), # Responsable
                "Clave" if getattr(r, 'hito_critico', False) else "No Clave",
                estado_calculado,
                r.fecha_limite.strftime("%d/%m/%Y") if r.fecha_limite else "",
                r.hora_limite.strftime("%H:%M") if r.hora_limite else "",
                r.fecha_estado.strftime("%d/%m/%Y") if r.fecha_estado else "", # Fecha y Hora Actualización
                gestor,
                observaciones,
                # Columnas extra sin Obligatorio
                fecha_cumplimiento,
                hora_cumplimiento,
                fecha_creacion,
                dept_cumplimiento
            ])

            # Aplicar color
            fila_numero = ws.max_row
            fill_color = colores_estado.get(estado_calculado)
            if fill_color:
                for cell in ws[fila_numero]:
                    cell.fill = fill_color
                    cell.font = font_blanco

        # Auto-ajustar columnas
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                # Ignorar filas de filtro para el ajuste de ancho para evitar columnas gigantes por los titulos
                if cell.row < 15:
                    continue
                try:
                    val = str(cell.value)
                    if len(val) > max_length:
                        max_length = len(val)
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
